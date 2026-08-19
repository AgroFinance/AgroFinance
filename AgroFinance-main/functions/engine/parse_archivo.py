"""AgroFinance — Puerto a Python de src/lib/parsing/parseArchivo.ts

Convierte un archivo (.xlsx/.csv ya soportados; .xml UBL 2.1 soportado
por separado en parsear_ubl) en LineaLeida[], el mismo formato que
consume ghg_classify. Corre en la Cloud Function, sobre el archivo ya
descargado de Cloud Storage — a diferencia del cliente, aquí no hay
límite de tiempo de un hilo de UI, pero se mantienen las mismas reglas:
una hoja/fila oculta se marca, no se descarta en silencio; un archivo
ilegible levanta un error con motivo, nunca se traga en silencio.

Alcance de este puerto: .xlsx, .csv y .xml (SUNAT UBL 2.1) — los tres
formatos que de hecho llegan hoy por Configuración/Analizar Datos y
por /upload (que es sobre todo un lector de facturas XML). .xls
(binario legado) y .ods quedan fuera por ahora — openpyxl no los lee —
y levantan ErrorArchivo con un mensaje explícito en vez de fingir
soporte.
"""

from __future__ import annotations

import csv
import math
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from .ghg_classify import LineaLeida


class ErrorArchivo(Exception):
    pass


# ------------------------------------------------------------
# Deteccion de unidad por sufijo del nombre de columna — mismos
# patrones que SUFIJOS_UNIDAD en parseArchivo.ts, en el mismo orden.
# ------------------------------------------------------------
_SUFIJOS_UNIDAD: list[tuple[re.Pattern, str]] = [
    (re.compile(r"_?(kwh)$", re.I), "kWh"),
    (re.compile(r"_?(mwh)$", re.I), "MWh"),
    (re.compile(r"_?(gal|galones)$", re.I), "gal"),
    (re.compile(r"_?(lt|ltr|litros?|l)$", re.I), "L"),
    (re.compile(r"_?(kg|kilos)$", re.I), "kg"),
    (re.compile(r"_?(tn|ton|toneladas?|t)$", re.I), "t"),
    (re.compile(r"_?(km)$", re.I), "km"),
    (re.compile(r"_?(u|und|unidades?)$", re.I), "u"),
    (re.compile(r"_?(pct|porcentaje|%)$", re.I), "%"),
]


def unidad_de_columna(col: str) -> str:
    for patron, unidad in _SUFIJOS_UNIDAD:
        if patron.search(col):
            return unidad
    return ""


def _a_numero(v) -> Optional[float]:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v) if math.isfinite(v) else None
    if not isinstance(v, str):
        return None
    limpio = v.replace(" ", "").replace(",", ".")
    if limpio == "":
        return None
    try:
        return float(limpio)
    except ValueError:
        return None


@dataclass
class ResultadoParseo:
    lineas: list[LineaLeida]
    hojas: list[str]
    columnas: list[str]
    filas_preview: list[list]


# ------------------------------------------------------------
# XLSX vía openpyxl
# ------------------------------------------------------------
def _parsear_xlsx(ruta: Path) -> ResultadoParseo:
    from openpyxl import load_workbook

    wb = load_workbook(ruta, data_only=True)
    if not wb.sheetnames:
        raise ErrorArchivo("El libro no contiene hojas legibles")

    lineas: list[LineaLeida] = []
    hojas: list[str] = []
    columnas: list[str] = []
    filas_preview: list[list] = []

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        hoja_oculta = ws.sheet_state != "visible"
        hojas.append(f"{nombre_hoja} (oculta)" if hoja_oculta else nombre_hoja)

        filas_iter = ws.iter_rows(values_only=False)
        try:
            encabezado_row = next(filas_iter)
        except StopIteration:
            continue
        cols = [str(c.value) if c.value is not None else "" for c in encabezado_row]
        if not any(cols):
            continue

        primera_hoja_con_datos = not columnas
        if primera_hoja_con_datos:
            columnas = cols

        for i, fila in enumerate(filas_iter, start=2):
            fila_oculta = hoja_oculta or bool(ws.row_dimensions[i].hidden)
            valores_fila = [c.value for c in fila]

            if primera_hoja_con_datos and len(filas_preview) < 12:
                filas_preview.append([v if v is not None else "" for v in valores_fila])

            for col_nombre, valor_crudo in zip(cols, valores_fila):
                if not col_nombre:
                    continue
                valor = _a_numero(valor_crudo)
                if valor is None:
                    continue
                lineas.append(LineaLeida(
                    id=f"{nombre_hoja}!{i}:{col_nombre}",
                    campo_leido=col_nombre,
                    valor=valor,
                    unidad=unidad_de_columna(col_nombre),
                    hoja=nombre_hoja,
                    fila=i,
                    oculto=fila_oculta or None,
                ))

    if not lineas:
        raise ErrorArchivo("No se encontró ninguna columna numérica que pueda leerse como consumo")

    return ResultadoParseo(lineas, hojas, columnas, filas_preview)


# ------------------------------------------------------------
# CSV vía csv estándar (misma lógica de columnas que el XLSX,
# sin concepto de hoja/fila oculta)
# ------------------------------------------------------------
def _parsear_csv(ruta: Path) -> ResultadoParseo:
    with open(ruta, newline="", encoding="utf-8-sig") as fh:
        lector = csv.reader(fh)
        try:
            cols = next(lector)
        except StopIteration:
            raise ErrorArchivo("El archivo no contiene hojas legibles")

        lineas: list[LineaLeida] = []
        filas_preview: list[list] = []

        for i, fila in enumerate(lector, start=2):
            if len(filas_preview) < 12:
                filas_preview.append(list(fila))
            for col_nombre, valor_crudo in zip(cols, fila):
                if not col_nombre:
                    continue
                valor = _a_numero(valor_crudo)
                if valor is None:
                    continue
                lineas.append(LineaLeida(
                    id=f"CSV!{i}:{col_nombre}",
                    campo_leido=col_nombre,
                    valor=valor,
                    unidad=unidad_de_columna(col_nombre),
                    hoja="CSV",
                    fila=i,
                ))

    if not lineas:
        raise ErrorArchivo("No se encontró ninguna columna numérica que pueda leerse como consumo")

    return ResultadoParseo(lineas, ["CSV"], cols, filas_preview)


# ------------------------------------------------------------
# XML SUNAT UBL 2.1 — mismos namespaces y misma lógica que parsearUBL
# en parseArchivo.ts (getElementsByTagName ahí == .iter() acá: ambos
# buscan en TODO el subárbol, no solo hijos directos).
# ------------------------------------------------------------
_NS = {
    "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
    "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
}

_UNIDAD_UBL = {"LTR": "L", "GLL": "gal", "KWH": "kWh", "MWH": "MWh", "KGM": "kg", "TNE": "t", "NIU": "u", "ZZ": ""}


def _tag(prefix: str, local: str) -> str:
    return f"{{{_NS[prefix]}}}{local}"


def _primero(elem: ET.Element, prefix: str, local: str) -> Optional[ET.Element]:
    return next(elem.iter(_tag(prefix, local)), None)


def _texto(elem: ET.Element, prefix: str, local: str) -> str:
    nodo = _primero(elem, prefix, local)
    return (nodo.text or "").strip() if nodo is not None else ""


@dataclass
class CabeceraUBL:
    invoice_id: str
    ruc: str
    proveedor: str
    fecha: str
    moneda: str
    monto_total: Optional[float]


@dataclass
class ResultadoUBL(ResultadoParseo):
    cabecera: CabeceraUBL = None  # type: ignore[assignment]


def _parsear_ubl(ruta: Path, nombre: str) -> ResultadoUBL:
    try:
        root = ET.parse(ruta).getroot()
    except ET.ParseError as exc:
        raise ErrorArchivo(f"El XML no está bien formado y no pudo interpretarse: {exc}")

    lineas_xml = list(root.iter(_tag("cac", "InvoiceLine")))
    raiz_ids = list(root.iter(_tag("cbc", "ID")))

    pago = _primero(root, "cbc", "PayableAmount")
    cabecera = CabeceraUBL(
        invoice_id=(raiz_ids[0].text.strip() if raiz_ids and raiz_ids[0].text else nombre),
        ruc=(raiz_ids[1].text.strip() if len(raiz_ids) > 1 and raiz_ids[1].text else "—"),
        proveedor=_texto(root, "cbc", "RegistrationName") or "—",
        fecha=_texto(root, "cbc", "IssueDate") or "—",
        moneda=(pago.get("currencyID") if pago is not None and pago.get("currencyID") else "PEN"),
        monto_total=_a_numero(_texto(root, "cbc", "PayableAmount")),
    )

    lineas: list[LineaLeida] = []
    filas_preview: list[list] = []
    for i, nodo in enumerate(lineas_xml):
        qty_nodo = _primero(nodo, "cbc", "InvoicedQuantity")
        descripcion = _texto(nodo, "cbc", "Description") or f"Ítem {i + 1}"
        unit_code = qty_nodo.get("unitCode", "") if qty_nodo is not None else ""
        valor = _a_numero(qty_nodo.text) if qty_nodo is not None else None
        unidad = _UNIDAD_UBL.get(unit_code, unit_code)
        lineas.append(LineaLeida(
            id=f"UBL!linea-{i + 1}", campo_leido=descripcion, valor=valor, unidad=unidad,
            hoja="Comprobante UBL 2.1", fila=i + 1,
            crudo=(qty_nodo.text.strip() if qty_nodo is not None and qty_nodo.text else None),
        ))
        if len(filas_preview) < 12:
            filas_preview.append([descripcion, valor if valor is not None else "", unidad])

    if not lineas:
        raise ErrorArchivo("El comprobante no declara líneas de detalle (cac:InvoiceLine)")

    return ResultadoUBL(
        lineas=lineas, hojas=["Comprobante UBL 2.1"],
        columnas=["Descripción del ítem", "Cantidad", "Unidad"], filas_preview=filas_preview,
        cabecera=cabecera,
    )


# ------------------------------------------------------------
# Punto de entrada
# ------------------------------------------------------------
def parsear_archivo(ruta_local: str, tipo: str) -> ResultadoParseo:
    ruta = Path(ruta_local)
    tipo = tipo.lower().lstrip(".")

    if tipo == "xlsx":
        return _parsear_xlsx(ruta)
    if tipo == "csv":
        return _parsear_csv(ruta)
    if tipo == "xml":
        return _parsear_ubl(ruta, ruta.name)
    if tipo in ("xls", "ods"):
        raise ErrorArchivo(
            f"Formato .{tipo} aún no soportado en el procesamiento en la nube — "
            "usa .xlsx, .csv o .xml, o súbelo desde Configuración (que sí lo lee en el navegador)."
        )
    raise ErrorArchivo(f"Formato .{tipo} no soportado")
